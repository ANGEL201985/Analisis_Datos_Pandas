
import numpy as np 
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, Protection, Alignment
import copy


x = pd.read_excel('Calidad1.xlsx')

# Especifica la columna y el rango de filas
columna = 'RESISTENCIA TESTIGO'
primera_fila = 2
ultima_fila = 1189

columna_filtrada = x.loc[primera_fila-1:ultima_fila-1, columna]


# Elimina los valores cero o vacíos de la columna seleccionada usando la funcion "dropna()" filtramos los valores que sean diferentes de cero y con el metodo "dropna()" eliminamos los ceros
columna_filtrada_sin_ceros = columna_filtrada[columna_filtrada !=0].dropna()

# Calculamos el promedio con todos los elementos de dicha columna usando la funcion de numpy "mean()"
promedio = columna_filtrada_sin_ceros.mean()

# A la tabla incial almacenada en x en las celdas vacias de la columna RESISTENCIA TESTIGO le agregaremos el valor de promedio de dicha columna que calculamos
x['RESISTENCIA TESTIGO'].fillna(promedio, inplace=True)

# Guardar el DataFrame modificado en un nuevo archivo Excel con Pandas
x.to_excel('Calidad_Nuevo.xlsx', index=False)


# Cargar el archivo original con openpyxl
original_wb = load_workbook('Calidad1.xlsx')
original_ws = original_wb.active

# Cargar el nuevo archivo creado por Pandas con openpyxl
nuevo_wb = load_workbook('Calidad_Nuevo.xlsx')
nuevo_ws = nuevo_wb.active

# Copiar las fórmulas del archivo original al nuevo archivo
#Iteramos sobre cada fila de la hoja de calculo original, iter_rows() es un método de openpyxl que permite iterar sobre las filas de una hoja de cálculo dentro de un rango específico.
for filas in original_ws.iter_rows(min_row = 1, max_row = original_ws.max_row, min_col = 1, max_col = original_ws.max_column):
    for celda in filas:
        nueva_celda = nuevo_ws[celda.coordinate]

        if celda.data_type == 'f':
            nueva_celda.value = celda.value

# Copiar los estilos de celdas desde el archivo original al nuevo archivo
for filas in original_ws.iter_rows(min_row = 1, max_row = original_ws.max_row, min_col = 1, max_col = original_ws.max_column):
    for celda in filas:
        nueva_celda = nuevo_ws[get_column_letter(celda.column) + str(celda.row)]
        nueva_celda.font = copy.copy(celda.font) if celda.font else Font()
        nueva_celda.border = copy.copy(celda.border) if celda.border else Border()
        nueva_celda.fill = copy.copy(celda.fill) if celda.fill else PatternFill()
        nueva_celda.alignment = copy.copy(celda.alignment) if celda.alignment else Alignment()
        nueva_celda.protection = copy.copy(celda.protection) if celda.protection else Protection()
        nueva_celda.number_format = celda.number_format

nuevo_wb.save('Calidad_Nuevo.xlsx')








